import os
import torch
import re
from comfy.sd1_clip import gen_empty_tokens
import random
from pathlib import Path
import folder_paths
import comfy
import os, re, io, base64, csv, shutil, requests, chardet, pathlib
import numpy as np
from pathlib import Path

from PIL import Image as PILImage
from io import BytesIO



from ..main_unit import *


#------------------------------------------------------------
# 安全导入检查 -- 将导入语句修改为以下形式
try:
    from transformers import T5Tokenizer, T5ForConditionalGeneration
except ImportError:
    T5Tokenizer = None
    T5ForConditionalGeneration = None
    print("Warning: transformers not installed, SuperPrompter node will not be available")

try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("Warning: openpyxl not installed, Excel-related nodes will not be available")

try:
    from openpyxl.drawing.image import Image as OpenpyxlImage
except ImportError:
    OpenpyxlImage = None
    print("Warning: openpyxl.drawing.image not available")

try:
    from openpyxl.utils import get_column_letter
except ImportError:
    get_column_letter = None
    print("Warning: openpyxl.utils.get_column_letter not available")

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None
    print("Warning: openpyxl.Workbook not available")

#------------------------------------------------------------


#region----------------------------------------------------------------------#

class text_SuperPrompter:
    def __init__(self):
        self.modelDir = os.path.expanduser("~") + "/.models"
        self.tokenizer = None
        self.model = None

    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "prompt": ("STRING", {"default": "Enter prompt here"}),
                "max_new_tokens": ("INT", {"default": 77, "min": 1, "max": 2048}),
                "repetition_penalty": ("FLOAT", {"default": 1.2, "min": 0.0, "max": 2.0, "step": 0.1}),
                "remove_incomplete_sentences": ("BOOLEAN", {"default": True}),
            },
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("generated_text",)
    FUNCTION = "generate_text"
    CATEGORY = "Apt_Preset/prompt"

    def remove_incomplete_sentence(self, paragraph):
        return re.sub(r'((?:\[^.!?\](?!\[.!?\]))\*+\[^.!?\\s\]\[^.!?\]\*$)', '', paragraph.rstrip())

    def download_models(self):
        model_name = "roborovski/superprompt-v1"
        self.tokenizer = T5Tokenizer.from_pretrained(model_name)
        self.model = T5ForConditionalGeneration.from_pretrained(model_name, torch_dtype=torch.float16)
        os.makedirs(self.modelDir, exist_ok=True)
        self.tokenizer.save_pretrained(self.modelDir)
        self.model.save_pretrained(self.modelDir)
        print("Downloaded SuperPrompt-v1 model files to", self.modelDir)

    def load_models(self):
        if not all(os.path.exists(self.modelDir) for file in self.modelDir):
            self.download_models()
        else:
            print("Model files found. Skipping download.")

        self.tokenizer = T5Tokenizer.from_pretrained(self.modelDir)
        self.model = T5ForConditionalGeneration.from_pretrained(self.modelDir, torch_dtype=torch.float16)
        print("SuperPrompt-v1 model loaded successfully.")

    def generate_text(self, prompt, max_new_tokens, repetition_penalty, remove_incomplete_sentences):
        if self.tokenizer is None or self.model is None:
            self.load_models()

        seed = 1
        torch.manual_seed(seed)
        device = 'cuda' if torch.cuda.is_available() else 'cpu'
        systemprompt = "Expand the following prompt to add more detail:"
        input_ids = self.tokenizer(systemprompt + prompt, return_tensors="pt").input_ids.to(device)
        if torch.cuda.is_available():
            self.model.to('cuda')

        outputs = self.model.generate(input_ids, max_new_tokens=max_new_tokens, repetition_penalty=repetition_penalty,
                                      do_sample=True)

        dirty_text = self.tokenizer.decode(outputs[0])
        text = dirty_text.replace("<pad>", "").replace("</s>", "").strip()
        
        if remove_incomplete_sentences:
            text = self.remove_incomplete_sentence(text)
        
        return (text,)



class text_mul_replace:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "text": ("STRING", {"multiline": True}),
                "target": ("STRING", {
                    "multiline": False,
                    "default": "man, dog "
                }),
                "replace": ("STRING", {
                    "multiline": False,
                    "default": "dog, man, "
                })
            }
        }
        
    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("text",)
    FUNCTION = "replace"
    CATEGORY = "Apt_Preset/prompt"

    def replace(self, text, target, replace):
        def split_with_quotes(s):
            pattern = r'"([^"]*)"|\s*([^,]+)'
            matches = re.finditer(pattern, s)
            return [match.group(1) or match.group(2).strip() for match in matches if match.group(1) or match.group(2).strip()]
        
        targets = split_with_quotes(target)
        exchanges = split_with_quotes(replace)
        

        word_map = {}
        for target, exchange in zip(targets, exchanges):

            target_clean = target.strip('"').strip().lower()
            exchange_clean = exchange.strip('"').strip()
            word_map[target_clean] = exchange_clean
        

        sorted_targets = sorted(word_map.keys(), key=len, reverse=True)
        
        result = text
        for target in sorted_targets:
            if ' ' in target:
                pattern = re.escape(target)
            else:
                pattern = r'\b' + re.escape(target) + r'\b'
            
            result = re.sub(pattern, word_map[target], result, flags=re.IGNORECASE)
        
        
        return (result,)

    @classmethod
    def IS_CHANGED(cls, text, target, replace):
        return (text, target, replace)



class text_mul_remove:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "text": ("STRING", {"multiline": True}),
                "words_to_remove": ("STRING", {
                    "multiline": False,
                    "default": "man, woman, world"
                })
            }
        }


    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("text",)
    FUNCTION = "clean_prompt"
    CATEGORY = "Apt_Preset/prompt"

    def clean_prompt(self, text, words_to_remove):
        remove_words = [word.strip().lower() for word in words_to_remove.split(',')]
        words = re.findall(r'\b\w+\b|[^\w\s]', text)
        
        cleaned_words = []
        skip_next = False
        for i, word in enumerate(words):
            word_lower = word.lower()
            if word_lower in remove_words:
                continue
            cleaned_words.append(word)
        cleaned_text = ' '.join(cleaned_words)

        return (cleaned_text,)

    @classmethod
    def IS_CHANGED(cls, text, words_to_remove):
        return (text, words_to_remove)



class text_free_wildcards:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "prompt": ("STRING", {"multiline": True}),
                "seed": ("INT", {
                    "default": 0,
                    "min": 0,
                    "max": 0xFFFFFFFF
                }),
                "wildcard_symbol": ("STRING", {"default": "@@"}),
                "recursive_depth": ("INT", {
                    "default": 5,
                    "min": 1,
                    "max": 10,
                    "step": 1
                })
            }
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("text",)
    FUNCTION = "process_wildcards"
    CATEGORY = "Apt_Preset/prompt"

    def process_wildcards(self, prompt, seed, wildcard_symbol, recursive_depth):
        random.seed(seed)
        wildcards_folder = Path(__file__).parent.parent  / "wildcards"
        
        logger.debug(f"Wildcards folder: {wildcards_folder}")
        logger.debug(f"Current working directory: {os.getcwd()}")
        logger.debug(f"Directory contents of wildcards folder: {os.listdir(wildcards_folder)}")
        
        def replace_wildcard(match, depth=0):
            if depth >= recursive_depth:
                logger.debug(f"Max depth reached: {depth}")
                return match.group(0)
            
            wildcard = match.group(1)
            file_path = os.path.join(wildcards_folder, f"{wildcard}.txt")
            logger.debug(f"Looking for file: {file_path} (depth: {depth})")
            logger.debug(f"File exists: {os.path.exists(file_path)}")
            
            if os.path.exists(file_path):
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        lines = [line.strip() for line in f if line.strip()]
                    if lines:
                        choice = random.choice(lines)
                        logger.debug(f"Replaced {wildcard} with: {choice} (depth: {depth})")
                        
                        if wildcard_symbol in choice:
                            logger.debug(f"Found nested wildcard in: {choice}")
                            processed_choice = re.sub(pattern, lambda m: replace_wildcard(m, depth + 1), choice)
                            logger.debug(f"After recursive processing: {processed_choice} (depth: {depth})")
                            return processed_choice
                        else:
                            return choice
                    else:
                        logger.warning(f"File {file_path} is empty")
                        return match.group(0)
                except Exception as e:
                    logger.error(f"Error reading file {file_path}: {str(e)}")
                    return match.group(0)
            else:
                logger.warning(f"File not found: {file_path}")
                return match.group(0)

        escaped_symbol = re.escape(wildcard_symbol)
        pattern = f"{escaped_symbol}([a-zA-Z0-9_]+)"
        

        
        processed_prompt = prompt
        for i in range(recursive_depth):
            new_prompt = re.sub(pattern, lambda m: replace_wildcard(m, 0), processed_prompt)
            if new_prompt == processed_prompt:
                break
            processed_prompt = new_prompt
            logger.debug(f"Iteration {i+1} result: {processed_prompt}")
        
        logger.debug(f"Final processed prompt: {processed_prompt}")
        
        return (processed_prompt,)

    @classmethod
    def IS_CHANGED(cls, prompt, seed, wildcard_symbol, recursive_depth):
        return float(seed)




#region---------------------------# Wildcards-------------


wildcards_dir1 = Path(__file__).parent.parent  / "wildcards"
os.makedirs(wildcards_dir1, exist_ok=True)
wildcards_dir2 = Path(folder_paths.base_path) / "wildcards"
# os.makedirs(wildcards_dir2, exist_ok=True)
print(f"Using wildcards dir:{wildcards_dir1} or {wildcards_dir2}")

full_dirs = [wildcards_dir1, wildcards_dir2]

WILDCARDS_LIST = (
    ["None"]
    + [
        "dir1 | " + str(wildcard.relative_to(wildcards_dir1))[:-4]
        for wildcard in wildcards_dir1.rglob("*.txt")
    ]
    + [
        "base_path | " + str(wildcard.relative_to(wildcards_dir2))[:-4]
        for wildcard in wildcards_dir2.rglob("*.txt")
    ]
)


class text_stack_wildcards:
    @classmethod
    def INPUT_TYPES(s):
        inputs = {
            "required": {
                "wildcards_count": (
                    "INT",
                    {"default": 1, "min": 1, "max": 50, "step": 1},
                ),
                "seed": ("INT", {"default": 0, "min": 0, "max": 0xFFFFFFFFFFFFFFFF}),
            },
            "optional": {
                "text": (
                    "STRING",
                    {
                        "default": "",
                        "multiline": True,
                    },
                ),
            },
        }

        for i in range(1, 10):
            inputs["required"][f"wildcard_name_{i}"] = (
                WILDCARDS_LIST,
                {"default": WILDCARDS_LIST[0]},
            )

        return inputs

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("text",)
    FUNCTION = "stack_Wildcards"
    CATEGORY = "Apt_Preset/prompt"

    def stack_Wildcards(self, wildcards_count, seed, text=None, **kwargs):

        selected_wildcards = [
            kwargs[f"wildcard_name_{i}"] for i in range(1, wildcards_count + 1)
        ]
        results = []

        for full_dir in full_dirs:
            for root, dirs, files in os.walk(full_dir):
                for wildcard in selected_wildcards:
                    if wildcard == "None":
                        continue
                    else:
                        if wildcard.startswith("dir1 | "):
                            wildcard_filename = wildcard[len("dir1 | ") :]
                            target_dir = wildcards_dir1
                        if wildcard.startswith("base_path | "):
                            wildcard_filename = wildcard[len("base_path | ") :]
                            target_dir = wildcards_dir2
                        if target_dir:
                            wildcard_file = (
                                Path(target_dir) / f"{wildcard_filename}.txt"
                            )
                            if wildcard_file.is_file():
                                with open(wildcard_file, "r", encoding="utf-8") as f:
                                    lines = f.readlines()
                                    if lines:
                                        selected_line_index = seed - 1
                                        selected_line_index %= len(lines)
                                        selected_line = lines[
                                            selected_line_index
                                        ].strip()
                                        results.append(selected_line)
                            else:
                                print(f"Wildcard File not found: {wildcard_file}")

                joined_result = ", ".join(results)
                if text == "":
                    joined_result = f"{joined_result}"
                else:
                    joined_result = f"{text},{joined_result}"
                return (joined_result,)


#endregion---------------------------# Wildcards-------------


class text_mul_Join:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                f"text{i+1}": ("STRING", {"default": "", "multiline": False}) for i in range(8)
            },
            "optional": {
                "delimiter": ("STRING", {
                    "default": "\\n",
                    "multiline": False,
                    "tooltip": "Use \\n for newline, \\t for tab, \\s for space"
                }),
            }
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("joined_text",)
    FUNCTION = "join_text"
    CATEGORY = "Apt_Preset/prompt"

    def join_text(self, delimiter, **kwargs):
        # 处理特殊转义字符
        if delimiter == "\\n":
            actual_delimiter = "\n"
        elif delimiter == "\\t":
            actual_delimiter = "\t"
        elif delimiter == "\\s":
            actual_delimiter = " "
        else:
            actual_delimiter = delimiter.strip()

        # 获取所有输入
        inputs = [kwargs[f"text{i+1}"] for i in range(8)]

        # 去除每个输入的首尾空白
        stripped_inputs = [text.strip() for text in inputs]

        # 找到最后一个非空索引
        last_non_empty_index = -1
        for i, text in enumerate(stripped_inputs):
            if text:
                last_non_empty_index = i

        # 构建结果
        result = []
        for i, text in enumerate(stripped_inputs):
            if i <= last_non_empty_index:
                result.append(text if text else "")  # 空字段也保留占位

        # 拼接并返回
        joined = actual_delimiter.join(result)
        return (joined,)


class text_mul_Split:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "text": ("STRING", {"multiline": True}),
                "delimiter": ("STRING", {
                    "default": "\\n",
                    "multiline": False,
                    "tooltip": "Use \\n for newline, \\t for tab, \\s for space"
                }),
            },
        }

    RETURN_TYPES = ("LIST", "STRING", "STRING", "STRING", "STRING", 
                    "STRING", "STRING", "STRING", "STRING")
    RETURN_NAMES = ("list_output", "item1", "item2", "item3", "item4", 
                    "item5", "item6", "item7", "item8")
    #OUTPUT_IS_LIST = (True, False, False, False, False, False, False, False, False)
    FUNCTION = "split_text"
    CATEGORY = "Apt_Preset/prompt"

    def split_text(self, text, delimiter):
        # 处理特殊转义字符
        if delimiter == "\\n":
            actual_delimiter = "\n"
        elif delimiter == "\\t":
            actual_delimiter = "\t"
        elif delimiter == "\\s":
            actual_delimiter = " "
        else:
            actual_delimiter = delimiter.strip()

        # 使用实际分隔符进行分割
        parts = [part.strip() for part in text.split(actual_delimiter)]

        # 生成8个固定输出，不足补空字符串
        output_items = parts[:8]
        while len(output_items) < 8:
            output_items.append("")


        list_out = []
        for text_item in parts:
            list_out.append(text_item)
        

        return (list_out, *output_items)


#endregion--------------------------------------


class excel_search_data:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "search_content": ("STRING", {"default": ""}),
                "search_mode": (["Precise_search", "Fuzzy_search"], {"default": "Precise_search"}),
            },
            "optional": {   } 
        }

    RETURN_TYPES = ("STRING", "INT", "INT")
    RETURN_NAMES = ("debug", "row", "col")
    FUNCTION = "search_data"
    CATEGORY = "Apt_Preset/prompt"

    def IS_CHANGED(): return float("NaN")

    def search_data(self, excel_path, sheet_name, search_content, search_mode):
        try:
            if not os.path.exists(excel_path):
                return (f"Error: File does not exist at path: {excel_path}", None, None)
            if not os.access(excel_path, os.R_OK):
                return (f"Error: No read permission for file at path: {excel_path}", None, None)
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            results = []
            found_row = None
            found_col = None
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell_value = cell.value if cell.value is not None else ""
                    cell_value_str = str(cell_value)
                    if (search_mode == "Precise_search" and cell_value_str == search_content) or \
                        (search_mode == "Fuzzy_search" and search_content in cell_value_str):
                        results.append(f"{sheet_name}|{row}|{col}|{cell_value}")
                        found_row = row
                        found_col = col

            workbook.close()
            del workbook
            if not results:
                return ("No results found.", None, None)
            return ("\n".join(results), found_row, found_col)
        except Exception as e:
            return (f"Error: {str(e)}", None, None)


class excel_row_diff:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "col_data": ("INT", {"default": 1, "min": 1, "step": 1}),
                "col_finish": ("INT", {"default": 1, "min": 1, "step": 1}),
            },
            "optional": {} 
        }

    RETURN_TYPES = ("INT", "INT", "INT")
    RETURN_NAMES = ("count", "count_data", "count_finish")
    FUNCTION = "excel_row_diff"
    CATEGORY = "Apt_Preset/prompt"
    DESCRIPTION = """
    - col_data=2: 统计第2列,从上到下连续非空单元格总数count1
    - col_finish=3: 统计第3列,从上到下连续非空单元格总数count2
    - 输出未对齐的数量:count_data-count_finish
    """

    def IS_CHANGED(cls): 
        return float("NaN")

    def excel_row_diff(self, excel_path, sheet_name, col_data, col_finish):
        try:
            if not os.path.exists(excel_path):
                raise Exception(f"Error: File does not exist at path: {excel_path}")

            if not os.access(excel_path, os.R_OK):
                raise Exception(f"Error: No read permission for file at path: {excel_path}")

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            def count_cells(col_index):
                if col_index == 0:  # 跳过无效列
                    return 0
                count = 0
                for row in range(1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row, column=col_index).value
                    if cell_value is not None:
                        count += 1
                    else:
                        break
                return count

            count1 = count_cells(col_data)
            count2 = count_cells(col_finish)

            result = abs(count1 - count2)

            workbook.close()
            del workbook

            return (result, count1, count2)

        except Exception as e:
            raise Exception(f"Error: {str(e)}")


class excel_column_diff:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row_data": ("INT", {"default": 1, "min": 1, "step": 1}),
                "row_finish": ("INT", {"default": 1, "min": 1, "step": 1}),
            },
            "optional": {}
        }

    RETURN_TYPES = ("INT", "INT", "INT")
    RETURN_NAMES = ("count", "count_data", "count_finish")
    FUNCTION = "excel_column_diff"
    CATEGORY = "Apt_Preset/prompt"
    DESCRIPTION = """
    - row_data=2: 统计第2行,从左到右连续非空单元格总数count1
    - row_finish=3: 统计第3行,从左到右连续非空单元格总数count2
    - 输出未对齐的数量:count_data-count_finish
    """

    def IS_CHANGED(cls):
        return float("NaN")

    def excel_column_diff(self, excel_path, sheet_name, row_data, row_finish):
        try:
            if not os.path.exists(excel_path):
                raise Exception(f"Error: File does not exist at path: {excel_path}")

            if not os.access(excel_path, os.R_OK):
                raise Exception(f"Error: No read permission for file at path: {excel_path}")

            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            def count_cells(row_index):
                if row_index == 0:  # 跳过无效行
                    return 0
                count = 0
                for col in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=row_index, column=col).value
                    if cell_value is not None:
                        count += 1
                    else:
                        break
                return count

            count1 = count_cells(row_data)
            count2 = count_cells(row_finish)

            result = abs(count1 - count2)

            workbook.close()
            del workbook

            return (result, count1, count2)

        except Exception as e:
            raise Exception(f"Error: {str(e)}")


class excel_read:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row_start": ("INT", {"default": 0, "min": 0}),
                "row_end": ("INT", {"default": 3, "min": 1}),
                "col_start": ("INT", {"default": 0, "min": 0}),
                "col_end": ("INT", {"default": 4, "min": 1}),
            },
            "optional": {}
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("data",)
    FUNCTION = "excel_read"
    CATEGORY = "Apt_Preset/prompt"
    DESCRIPTION = """
    - #excle最小单元行或列不能为0,将无效跳过
    - row_start=0, row_end=3: 单行输出: row=3    
    - row_start=1, row_end=3: 多行输出: row=1,2,3
    - row_start = row_end=1 : 行数相同,单行输出row=1
    - row_end=3, row_end=1 : 报错, 起始行必须小于=结束行
    """
    @classmethod
    def IS_CHANGED(cls):
        return float("NaN")

    def excel_read(self, excel_path, sheet_name, row_start, row_end, col_start, col_end):
        try:
            # 校验 start <= end 且 >= 0
            if row_start > row_end:
                raise Exception(f"Error: row_start ({row_start}) must be <= row_end ({row_end})!")
            if col_start > col_end:
                raise Exception(f"Error: col_start ({col_start}) must be <= col_end ({col_end})!")

            # 处理 row_start == 0 的情况：只取 row_end 行
            if row_start == 0:
                start_row = end_row = max(1, row_end)
            else:
                start_row, end_row = row_start, row_end

            # 处理 col_start == 0 的情况：只取 col_end 列
            if col_start == 0:
                start_col = end_col = max(1, col_end)
            else:
                start_col, end_col = col_start, col_end

            # 确保行列编号 ≥ 1
            start_row = max(1, start_row)
            end_row = max(1, end_row)
            start_col = max(1, start_col)
            end_col = max(1, end_col)

            # 打开 Excel 文件并读取数据
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook[sheet_name]

            output_lines = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                output_lines.append("|".join(row_data))

            workbook.close()
            del workbook

            return ("\n".join(output_lines),)

        except Exception as e:
            return (f"Error: {str(e)}",)


class excel_write_data:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row_start": ("INT", {"default": 0, "min": 0}),
                "row_end": ("INT", {"default": 3, "min": 1}),
                "col_start": ("INT", {"default": 0, "min": 0}),
                "col_end": ("INT", {"default": 5, "min": 1}),
                "data": ("STRING", {"default": "", "multiline": True}),
            },
            "optional": {}
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("debug",)
    FUNCTION = "write_data"
    CATEGORY = "Apt_Preset/prompt"
    DESCRIPTION = """
    - 示例: data 输入如下数据
    - 1 | 2 | 3 | 4
    - a | b | c | d
    - row_start=2, row_end=3  
    - col_start=2, col_end=5
    - 数据会依次填入第2行第2列 到 第3行第5列
    - 
    - row_start=0, row_end=3  
    - col_start=0, col_end=5
    - 数据会只填入到第3行第5列的单元格
    """

    @classmethod
    def IS_CHANGED(cls):
        return float("NaN")

    def write_data(self, excel_path, sheet_name, row_start, row_end, col_start, col_end, data):
        try:
            # 校验 start <= end 且 >= 0
            if row_start > row_end:
                return (f"Error: row_start ({row_start}) must be <= row_end ({row_end})",)
            if col_start > col_end:
                return (f"Error: col_start ({col_start}) must be <= col_end ({col_end})",)

            # 处理 start == 0 的情况：只写 end 行/列
            if row_start == 0:
                start_row = end_row = max(1, row_end)
            else:
                start_row, end_row = row_start, row_end

            if col_start == 0:
                start_col = end_col = max(1, col_end)
            else:
                start_col, end_col = col_start, col_end

            # 确保最小值为1（兼容 Excel）
            start_row = max(1, start_row)
            end_row = max(1, end_row)
            start_col = max(1, start_col)
            end_col = max(1, end_col)

            # 文件存在性及权限检查
            if not os.path.exists(excel_path):
                return (f"Error: File does not exist at path: {excel_path}",)
            if not os.access(excel_path, os.W_OK):
                return (f"Error: No write permission for file at path: {excel_path}",)

            # 加载工作簿和工作表
            workbook = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
            sheet = workbook[sheet_name]

            # 解析输入数据
            data_lines = data.strip().split("\n")

            # 写入数据
            for row_index, line in enumerate(data_lines, start=start_row):
                if row_index > end_row:
                    break
                cell_values = line.split("|")
                for col_index, cell_value in enumerate(cell_values, start=start_col):
                    if col_index > end_col:
                        break
                    if cell_value.strip():
                        sheet.cell(row=row_index, column=col_index).value = cell_value.strip()

            # 保存并关闭工作簿
            workbook.save(excel_path)
            workbook.close()
            del workbook

            return ("Data written successfully!",)

        except PermissionError as pe:
            return (f"Permission Error: {str(pe)}",)
        except Exception as e:
            return (f"Error: {str(e)}",)


class excel_insert_image:
    @classmethod
    def INPUT_TYPES(cls):
        return {
            "required": {
                "excel_path": ("STRING", {"default": "excel_file_path.xlsx"}),
                "sheet_name": ("STRING", {"default": "Sheet1"}),
                "row_start": ("INT", {"default": 0, "min": 0}),
                "row_end": ("INT", {"default": 1, "min": 1}),
                "col_start": ("INT", {"default": 0, "min": 0}),
                "col_end": ("INT", {"default": 1, "min": 1}),
                "img_height": ("INT", {"default": 256, "max": 2048, "min": 64}),
                "image_path": ("STRING", {"default": "image_file_path.png"}),
            },
        }

    RETURN_TYPES = ("STRING",)
    RETURN_NAMES = ("debug",)
    FUNCTION = "write_image"
    CATEGORY = "Apt_Preset/prompt"
    DESCRIPTION = """
    - 示例: 图片输入如下
    - row_start=2, row_end=3  
    - col_start=2, col_end=5
    - 图片会依次填入第2行第2列 到 第3行第5列
    - 
    - row_start=0, row_end=3  
    - col_start=0, col_end=5
    - 图片会只插入到第3行第5列的单元格
    """
    @classmethod
    def IS_CHANGED(cls):
        return float("NaN")

    def write_image(self, excel_path, sheet_name, row_start, row_end, col_start, col_end, image_path,img_height):
        try:
            # 校验 start <= end 且 >= 0
            if row_start > row_end:
                return (f"Error: row_start ({row_start}) must be <= row_end ({row_end})",)
            if col_start > col_end:
                return (f"Error: col_start ({col_start}) must be <= col_end ({col_end})",)

            # 处理 start == 0：只插入 end 所在行/列
            if row_start == 0:
                start_row = end_row = max(1, row_end)
            else:
                start_row, end_row = row_start, row_end

            if col_start == 0:
                start_col = end_col = max(1, col_end)
            else:
                start_col, end_col = col_start, col_end

            # 确保最小值为 1（兼容 Excel）
            start_row = max(1, start_row)
            end_row = max(1, end_row)
            start_col = max(1, start_col)
            end_col = max(1, end_col)

            # 文件存在性及权限检查
            if not os.path.exists(excel_path):
                return (f"Error: Excel file does not exist at path: {excel_path}",)
            if not os.access(excel_path, os.W_OK):
                return (f"Error: No write permission for Excel file at path: {excel_path}",)
            if not os.path.exists(image_path):
                return (f"Error: Image file does not exist at path: {image_path}",)
            if not os.access(image_path, os.R_OK):
                return (f"Error: No read permission for image file at path: {image_path}",)

            # 加载工作簿和工作表
            workbook = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
            sheet = workbook[sheet_name]

            # 插入图片的目标位置（仅使用 start 坐标）
            target_row = start_row
            target_col = start_col
            cell_address = get_column_letter(target_col) + str(target_row)

            # 打开图片并按比例缩放（高度固定为256像素，宽度按比例计算）
            with PILImage.open(image_path) as img:
                width, height = img.size
                scale = img_height / height
                target_width = int(width * scale)
                resized_img = img.resize((target_width, img_height), PILImage.LANCZOS)

                # 转换为字节流供 openpyxl 使用
                img_byte_arr = BytesIO()
                resized_img.save(img_byte_arr, format=img.format)
                openpyxl_img = OpenpyxlImage(img_byte_arr)

            # 调整单元格尺寸以适应图片
            column_letter = get_column_letter(target_col)
            
            # 设置列宽（按像素转磅的近似公式：1像素 ≈ 0.75磅）
            sheet.column_dimensions[column_letter].width = target_width * 0.75 / 7
            
            # 设置行高为192磅（对应256像素）
            sheet.row_dimensions[target_row].height = 192

            # 插入图片（图片会自动以单元格左上角对齐）
            sheet.add_image(openpyxl_img, cell_address)

            # 保存并关闭工作簿
            workbook.save(excel_path)
            workbook.close()

            return ("Image inserted and scaled successfully!",)

        except PermissionError as pe:
            return (f"Permission Error: {str(pe)}",)
        except Exception as e:
            return (f"Error: {str(e)}",)
        




class text_CSV_load:
    # 获取当前脚本所在目录
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    CSV_DIR = os.path.join(BASE_DIR, "CSV")

    @staticmethod
    def load_csv(csv_path: str):
        data = {"Error loading CSV, check the console": ["", ""]}
        if not os.path.exists(csv_path):
            return data
        try:
            with open(csv_path, "r", encoding="utf-8") as f:
                data = [[x.replace('"', '').replace('\n', '') for x in re.split(',(?=(?:[^"]*"[^"]*")*[^"]*$)', line)] for line in f.readlines()[1:]]
                data = {x[0]: [x[1], x[2]] for x in data}
        except Exception as e:
            print(f"Error loading CSV: {csv_path}. Error: {e}")
        return data

    @classmethod
    def INPUT_TYPES(cls):
        # 使用构建好的绝对路径加载 CSV 文件
        cls.artists_csv = cls.load_csv(os.path.join(cls.CSV_DIR, "A_artists.csv"))
        cls.material_csv = cls.load_csv(os.path.join(cls.CSV_DIR, "A_material.csv"))
        cls.nature_csv = cls.load_csv(os.path.join(cls.CSV_DIR, "E_nature.csv"))
        cls.scene_csv = cls.load_csv(os.path.join(cls.CSV_DIR, "E_scene.csv"))
        cls.building_csv = cls.load_csv(os.path.join(cls.CSV_DIR, "M_building.csv"))
        cls.cosplay_csv = cls.load_csv(os.path.join(cls.CSV_DIR, "M_cosplay.csv"))
        cls.camera_csv = cls.load_csv(os.path.join(cls.CSV_DIR, "T_camera.csv"))
        cls.cameraEffect_csv = cls.load_csv(os.path.join(cls.CSV_DIR, "T_cameraEffect.csv"))
        cls.detail_csv = cls.load_csv(os.path.join(cls.CSV_DIR, "T_detail.csv"))
        cls.light_csv = cls.load_csv(os.path.join(cls.CSV_DIR, "T_light.csv"))

        return {
            "required": {
                "A_artists": (list(cls.artists_csv.keys()),),
                "A_material": (list(cls.material_csv.keys()),),
                "M_cosplay": (list(cls.cosplay_csv.keys()),),
                "T_detail": (list(cls.detail_csv.keys()),),
                "T_light": (list(cls.light_csv.keys()),),
                "E_nature": (list(cls.nature_csv.keys()),),
                "E_scene": (list(cls.scene_csv.keys()),),
                "M_building1": (list(cls.building_csv.keys()),),
                "M_building2": (list(cls.building_csv.keys()),),
                "M_building3": (list(cls.building_csv.keys()),),
                "T_camera": (list(cls.camera_csv.keys()),),
                "T_cameraEffect": (list(cls.cameraEffect_csv.keys()),),
                "subject": ("STRING", {"multiline": True}),
            },
        }




    RETURN_TYPES = ("STRING", "STRING")
    RETURN_NAMES = ("pos", "neg")
    FUNCTION = "execute"
    CATEGORY = "Apt_Preset/prompt"

    def execute(self, A_artists, M_cosplay, T_detail, T_light, E_nature, E_scene, M_building1, M_building2, M_building3, A_material, T_camera, T_cameraEffect,subject=""):
        # Combine all positive and negative prompts
        positive_prompt = ""
        negative_prompt = ""

        # Aggregating all positive and negative prompts from the different CSV files
        for csv_dict in [self.artists_csv, self.building_csv, self.cosplay_csv, self.nature_csv, self.scene_csv,
                        self.material_csv, self.camera_csv, self.cameraEffect_csv, self.detail_csv, self.light_csv]:
            if A_artists in csv_dict:
                positive_prompt += csv_dict[A_artists][0] + " "
                negative_prompt += csv_dict[A_artists][1] + " "
            if M_cosplay in csv_dict:
                positive_prompt += csv_dict[M_cosplay][0] + " "
                negative_prompt += csv_dict[M_cosplay][1] + " "
            if T_detail in csv_dict:
                positive_prompt += csv_dict[T_detail][0] + " "
                negative_prompt += csv_dict[T_detail][1] + " "
            if T_light in csv_dict:
                positive_prompt += csv_dict[T_light][0] + " "
                negative_prompt += csv_dict[T_light][1] + " "
            if E_nature in csv_dict:
                positive_prompt += csv_dict[E_nature][0] + " "
                negative_prompt += csv_dict[E_nature][1] + " "
            if E_scene in csv_dict:
                positive_prompt += csv_dict[E_scene][0] + " "
                negative_prompt += csv_dict[E_scene][1] + " "
            if M_building1 in csv_dict:
                positive_prompt += csv_dict[M_building1][0] + " "
                negative_prompt += csv_dict[M_building1][1] + " "
            if M_building2 in csv_dict:
                positive_prompt += csv_dict[M_building2][0] + " "
                negative_prompt += csv_dict[M_building2][1] + " "
            if M_building3 in csv_dict:
                positive_prompt += csv_dict[M_building3][0] + " "
                negative_prompt += csv_dict[M_building3][1] + " "
            if A_material in csv_dict:
                positive_prompt += csv_dict[A_material][0] + " "
                negative_prompt += csv_dict[A_material][1] + " "
            if T_camera in csv_dict:
                positive_prompt += csv_dict[T_camera][0] + " "
                negative_prompt += csv_dict[T_camera][1] + " "
            if T_cameraEffect in csv_dict:
                positive_prompt += csv_dict[T_cameraEffect][0] + " "
                negative_prompt += csv_dict[T_cameraEffect][1] + " "

        pos = positive_prompt.strip()
        neg = negative_prompt.strip()

        prompt = "{prompt}"
        if prompt not in pos:
            pos = subject + " " + pos
        else:
            pos = pos.replace(prompt, subject)
        return (pos, neg)



class excel_Prompter:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    EXCEL_DIR = os.path.join(BASE_DIR, "excel")
    MAX_INPUTS = 15

    def __init__(self):
        self.excel_data = {}

    @staticmethod
    def extract_number(filename: str) -> int:
        """从文件名提取开头数字，用于排序"""
        match = re.match(r'^\d+', filename)
        return int(match.group(0)) if match else 0

    @staticmethod
    def load_excel(excel_path: str) -> dict:
        """加载Excel文件，返回{ID: [正向提示, 负向提示]}格式"""
        try:
            workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet = workbook.active
            headers = next(sheet.iter_rows(values_only=True))
            id_col = headers.index('ID') if 'ID' in headers else 0
            pos_col = headers.index('Positive') if 'Positive' in headers else 1
            neg_col = headers.index('Negative') if 'Negative' in headers else 2
            return {row[id_col]: [str(row[pos_col] or ''), str(row[neg_col] or '')] 
                    for row in sheet.iter_rows(min_row=2, values_only=True) if row[id_col]}
        except Exception as e:
            print(f"加载Excel失败: {excel_path} - {e}")
            return {"加载失败": ["", ""]}

    @classmethod
    def get_excel_files(cls) -> dict:
        """获取并排序Excel文件，限制最多加载MAX_INPUTS个"""
        if not os.path.exists(cls.EXCEL_DIR):
            print("Excel目录不存在")
            return {}
            
        files = [f for f in os.listdir(cls.EXCEL_DIR) if f.endswith('.xlsx')]
        
        if not files:
            print("目录中没有.xlsx文件")
            return {}
            
        files.sort(key=lambda x: (cls.extract_number(x), x))
        
        result = {}
        for f in files[:cls.MAX_INPUTS]:
            file_path = os.path.join(cls.EXCEL_DIR, f)
            file_data = cls.load_excel(file_path)
            result[f.replace('.xlsx', '')] = file_data
        
        return result

    @classmethod
    def INPUT_TYPES(cls) -> dict:
        """定义节点输入类型，为每个Excel文件创建下拉选项"""
        # 在每次调用INPUT_TYPES时都重新加载数据
        excel_data = cls.get_excel_files()
        
        inputs = {"required": {"subject": ("STRING", {"multiline": True})}}
        
        if not excel_data:
            # 如果没有Excel文件，提供默认选项
            for i in range(cls.MAX_INPUTS):
                inputs["required"][f"input_{i+1}"] = (["未找到Excel文件"],)
        else:
            # 为每个Excel文件创建一个下拉菜单
            file_types = list(excel_data.keys())
            for i in range(cls.MAX_INPUTS):
                if i < len(file_types):
                    file_type = file_types[i]
                    options = list(excel_data[file_type].keys())
                    if options:
                        inputs["required"][f"input_{i+1}"] = (options,)
                    else:
                        inputs["required"][f"input_{i+1}"] = (["无数据"],)
                else:
                    inputs["required"][f"input_{i+1}"] = (["无数据"],)
            
        return inputs

    RETURN_TYPES = ("STRING", "STRING")
    RETURN_NAMES = ("pos", "neg")
    FUNCTION = "execute"
    CATEGORY = "Apt_Preset/prompt"

    def execute(self, **kwargs) -> tuple:
        """组合所有选中的正负向提示词"""
        # 在执行时加载数据（如果尚未加载）
        if not self.excel_data:
            self.excel_data = self.get_excel_files()
        
        subject = kwargs.pop("subject", "")
        pos_prompt = ""
        neg_prompt = ""
        
        # 处理输入参数
        for param_name, param_value in kwargs.items():
            # 跳过提示信息等非输入参数
            if param_value in ["未找到Excel文件", "无数据", "加载失败"]:
                continue
                
            # 在所有数据中查找匹配的ID
            found = False
            for file_type, data in self.excel_data.items():
                if param_value in data:
                    pos_prompt += data[param_value][0] + " "
                    neg_prompt += data[param_value][1] + " "
                    found = True
                    break
            
            if not found and param_value not in ["请选择", "未找到Excel文件", "无数据", "加载失败"]:
                print(f"未找到ID: {param_value}")
        
        pos = pos_prompt.strip()
        neg = neg_prompt.strip()
        
        if "{prompt}" in pos:
            pos = pos.replace("{prompt}", subject)
        elif subject:
            pos = f"{subject} {pos}"
            
        return (pos, neg)



























